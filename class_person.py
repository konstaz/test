import os
from abc import ABC

from openpyxl import load_workbook



class Person(ABC):
    _name = None
    _last_name = None
    _phone = None
    _social_status = None

    def taxes(self):
        pass

    def social_status(self):
        pass

    def set_info(self):
        pass

    @property
    def phone_number(self):
        return self._phone

    @phone_number.setter
    def phone_number(self, value):
        if all([i.isdigit() for i in value]):
            self._phone = value
        else:
            raise ValueError('Phone number should contain only digits')

    @property
    def full_name(self):
        return f'{self._name} {self._last_name}'


    @full_name.setter
    def full_name(self, name, last_name):
        if all([i.isalpha() for i in name]):
            self._name = name
        else:
            raise ValueError('Name should contain only letters')

        if all([i.isalpha() for i in last_name]):
            self._last_name = last_name
        else:
            raise ValueError('Last name should contain only letters')

    @property
    def check_letters(self):
        pass

    @check_letters.setter
    def check_letters(self):
        pass


class Unemployed(Person):

    def __init__(self, name: str, last_name: str, phone: int):
        self._name = name
        self._last_name = last_name
        self._phone = phone
        self._social_status = "Unemployed"

        if all([i.isalpha() for i in name]):
            self._name = name
        else:
            raise ValueError('Name should contain only letters')

        if all([i.isalpha() for i in last_name]):
            self._last_name = last_name
        else:
            raise ValueError('Last name should contain only letters')

        if all([i.isdigit() for i in phone]):
            self._phone = phone
        else:
            raise ValueError('Phone number should contain only digits')


    @classmethod
    def unemployed(cls):
        return 'People of this do not have a job. Please assist and help them find a job', cls

    def taxes(self):
        return 0

    def social_status(self):
        return self._social_status

    @property
    def check_letters(self):
        return self._name, self._last_name

    @check_letters.setter
    def check_letters(self, name, last_name):
        if all([i.isalpha() for i in name]):
            self._name = name

        else:
            raise ValueError('Name should contain only letters')

        if all([i.isalpha() for i in last_name]):
            self._last_name = last_name
        else:
            raise ValueError('Last name should contain only letters')


class Worker:

    def __init__(self, name: str, last_name: str, phone, activity):  # , company: str, salary: float):
        self._name = name
        self._last_name = last_name
        self._phone = phone
        self._activity = activity
        self._company_name = None
        self._salary = None
        self._social_status = 'Worker'

        if all([i.isalpha() for i in name]):
            self._name = name

        else:
            raise ValueError('Name should contain only letters')

        if all([i.isalpha() for i in last_name]):
            self._last_name = last_name
        else:
            raise ValueError('Last name should contain only letters')

        # if all([i.isdigit() for i in phone]):
        #     self._phone = phone
        # else:
        #     raise TypeError('Phone number should contain only digits')

    def __str__(self):
        return f'Name={self._name}, Lastname={self._last_name}, phone={self._phone}, activity={self._activity}'

    def taxes(self):
        return self._salary * 0.2

    def social_status(self):
        return self._social_status

    @property
    def company_name(self):
        return self._company_name

    @company_name.setter
    def company_name(self, name: str):
        if all([i.isalpha() for i in name]):
            self._company_name = name
        else:
            raise TypeError('Company name should contain only letters not digits')

        self._company_name = name

    @property
    def salary(self):
        return self._salary

    @salary.setter
    def salary(self, salary):
        if salary < 0:
            raise ValueError('salary cannot be negative')
        elif salary.isalpha:
            raise ValueError('salary cannot contain letters')
        else:
            self._salary = salary

    @property
    def check_letters(self):
        return self._name, self._last_name

    @check_letters.setter
    def check_letters(self, name, last_name):
        if all([i.isalpha() for i in name]):
            self._name = name

        else:
            raise ValueError('Name should contain only letters')

        if all([i.isalpha() for i in last_name]):
            self._last_name = last_name
        else:
            raise ValueError('Last name should contain only letters')


class Programmer(Worker):
    prof = None

    def __new__(cls, *args, **kwargs):
        obj = super(Programmer, cls).__new__(cls)
        obj.prof = cls.prof
        return obj

    def __init__(self, name: str, last_name: str, phone, company: str, salary: int, nick: str):
        super().__init__(name, last_name, phone, company, salary)
        self._social_status = 'Programmer'
        self._nick = nick


    @classmethod
    def developer(cls):
        cls.prof = 'developer'
        return cls

    @classmethod
    def tester(cls):
        cls.prof = 'tester'
        return cls


    def taxes(self):
        return self._salary * 0.05

    def social_status(self):
        return self._social_status


# class SheetRepresentation(object):
#     excel_sheet_path = None
#     row_repr_class = None
#
#
#     def __str__(self):
#         workbook = load_workbook(self.excel_sheet_path)
#         return '\n'.join([self.row_repr_class(*i) for i in workbook.active.values[1:]])
#
#
# class PeopleProfession(ABC):
#     wb = load_workbook('profs.xlsx')
#     sheet = wb.active
#
#     def __init__(self, name, last_name, phone, social):
#         self._name = name
#         self._last_name = last_name
#         self._phone_number = phone
#         self._social_status = social
#
#     def __str__(self):
#         pass
#
#
# class ProfRepresentation(SheetRepresentation):
#     excel_sheet_path = 'profs.xlsx'
#     row_repr_class = PeopleProfession


class PeopleProfession:

    def __init__(self, file):

        self._datafile = file

        if not os.path.isfile(file):
            raise FileNotFoundError('file does not exist')

        self.wb = load_workbook(self._datafile)
        self.data = list(self.wb.active.values)[1:]


    def list_prof(self):
        return [Worker(i[0], i[1], i[2], i[3]) for i in self.data]


    def print_data(self):
        for i in self.data:
            worker = Worker(name=i[0], last_name=i[1], phone=i[2], activity=i[3])
            print(worker)



if __name__ == "__main__":
    x = PeopleProfession('profs.xlsx')

    print(x.list_prof())
    print(x.print_data())

    # z = PeopleProfession('prodsafsadffs.xlsx')

    # a = Programmer.developer()('alex', 'schulman', 888, 'top Perses', 30000, 'AlexDarkStalker2001')
    # b = Programmer.tester()('alex', 'schulman', 888, 'top Perses', 30000, 'AlexDarkStalker2001')
    # c = Programmer('alex', 'schulman', 888, 'top Perses', 30000, 'AlexDarkStalker2001')
    # print(a.prof)
    # print(b.prof)
    # print(c.prof)

    # x = Programmer('alex', 'schulman', 888, 'top Perses', 30000, 'AlexDarkStalker2001')

    # Programmer.programmer()
    # print(ProfRepresentation())